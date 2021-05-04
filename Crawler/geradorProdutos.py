import openpyxl

wb_obj = openpyxl.load_workbook('Casa_das_Cuecas.xlsx')

sheet_obj = wb_obj.active

# print(sheet_obj.max_row)
# print(sheet_obj.max_column)

#print(range(1, sheet_obj.max_row))

lista_produtos = ''

for i in range(2, sheet_obj.max_row):

    # Informações de produtos
    nome = sheet_obj.cell(row=i, column=3).value
    categoria = sheet_obj.cell(row=i, column=1).value
    img1 = sheet_obj.cell(row=i, column=8).value
    img2 = sheet_obj.cell(row=i, column=9).value
    img3 = sheet_obj.cell(row=i, column=10).value
    img4 = sheet_obj.cell(row=i, column=11).value
    img5 = sheet_obj.cell(row=i, column=12).value

    preco = sheet_obj.cell(row=i, column=6).value
    preco = "{:.2f}".format(float(preco))

    descricao_bruta = sheet_obj.cell(row=i, column=7).value.split('\n\n')[0]
    descricao = ''
    caracteres_pau = ['“', '”', "'", "''"]
    for letra in descricao_bruta:
        if letra not in caracteres_pau:
            descricao += letra

    sku = 7+i

    produto = f"{{nome: '{nome}', genero: '', categoria: '{categoria}',faixa: '',img1: '{img1}',img2: '{img2}',img3: '{img3}',img4: '{img4}',img5: '{img5}',preco: {preco},parcelado: '',descricao: '{str(descricao)}',loja: 'Casa das Cuecas',sku: {sku}, }},"

    lista_produtos += produto


f = open("produtos_casa_das_cuecas.txt", "a")
f.write(lista_produtos)
f.close()
